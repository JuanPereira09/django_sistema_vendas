from django.shortcuts import render
from django.db.models import Sum, Count
from django.utils.timezone import now
from .models import Venda, ItemVenda, Cliente, Produto
from django.db.models.functions import TruncMonth
from django.contrib.auth.decorators import login_required
from .forms import VendaForm, ItemVendaForm, ProdutoForm
from django.shortcuts import redirect
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

@login_required
def dashboard(request):
    mes_atual = now().month
    ano_atual = now().year

    vendas_mes = Venda.objects.filter(
        data__month=mes_atual,
        data__year=ano_atual
    )

    total_mes = vendas_mes.aggregate(total=Sum('valor_total'))['total'] or 0

    faturamento_total = Venda.objects.aggregate(
        total=Sum('valor_total')
    )['total'] or 0

    produto_mais_vendido = (
        ItemVenda.objects
        .values('produto__nome')
        .annotate(total_vendido=Sum('quantidade'))
        .order_by('-total_vendido')
        .first()
    )

    cliente_mais_comprou = (
        Cliente.objects
        .annotate(total_compras=Sum('venda__valor_total'))
        .order_by('-total_compras')
        .first()
    )

    vendas_por_mes = (
        Venda.objects
        .annotate(mes=TruncMonth('data'))
        .values('mes')
        .annotate(total=Sum('valor_total'))
        .order_by('mes')
)

    labels = []
    dados = []

    for venda in vendas_por_mes:
        labels.append(venda['mes'].strftime('%m/%Y'))
        dados.append(float(venda['total']))

    context = {
        'total_mes': total_mes,
        'faturamento_total': faturamento_total,
        'produto_mais_vendido': produto_mais_vendido,
        'cliente_mais_comprou': cliente_mais_comprou,
        'labels': labels,
        'dados': dados,
    }

    return render(request, 'dashboard.html', context)

@login_required
def adicionar_itens(request, venda_id):
    venda = get_object_or_404(Venda, id=venda_id)
    itens = venda.itens.all()

    if request.method == 'POST':
        form = ItemVendaForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.venda = venda
            item.preco_unitario = item.produto.preco
            item.save()
            return redirect('adicionar_itens', venda_id=venda.id)
    else:
        form = ItemVendaForm()

    return render(request, 'adicionar_itens.html', {
        'venda': venda,
        'form': form,
        'itens': itens
    })

@login_required
def nova_venda(request):
    if request.method == 'POST':
        form = VendaForm(request.POST)
        if form.is_valid():
            venda = form.save()
            return redirect('adicionar_itens', venda_id=venda.id)
    else:
        form = VendaForm()

    return render(request, 'nova_venda.html', {'form': form})

@login_required
def finalizar_venda(request, venda_id):
    venda = get_object_or_404(Venda, id=venda_id)

    return redirect('dashboard')

@login_required
def lista_produtos(request):
    produtos = Produto.objects.all()

    return render(request, 'produtos.html', {
        'produtos': produtos
    })

@login_required
def novo_produto(request):

    if request.method == 'POST':
        form = ProdutoForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('produtos')

    else:
        form = ProdutoForm()

    return render(request, 'novo_produto.html', {
        'form': form
    })

def pagina_relatorio(request):
    return render(request, 'relatorio.html')

def relatorio_vendas(request):

    data_inicio = request.GET.get("inicio")
    data_fim = request.GET.get("fim")

    itens = ItemVenda.objects.select_related('venda', 'produto')

    if data_inicio:
        itens = itens.filter(venda__data__date__gte=data_inicio)

    if data_fim:
        itens = itens.filter(venda__data__date__lte=data_fim)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio de Vendas"

    headers = [
        "Cliente",
        "Produto",
        "Quantidade",
        "Valor Unitario",
        "Total Item",
        "Data"
    ]

    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_geral = 0

    for item in itens:

        total_item = item.quantidade * item.produto.preco
        total_geral += total_item

        ws.append([
            item.venda.cliente.nome,
            item.produto.nome,
            item.quantidade,
            float(item.produto.preco),
            float(total_item),
            item.venda.data.strftime("%d/%m/%Y")
        ])

    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    ws.auto_filter.ref = ws.dimensions

    ws.append([])
    ws.append(["", "", "", "TOTAL GERAL:", float(total_geral)])

    total_cell = ws.cell(row=ws.max_row, column=4)
    total_cell.font = Font(bold=True)

    total_valor = ws.cell(row=ws.max_row, column=5)
    total_valor.number_format = 'R$ #,##0.00'
    total_valor.font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=relatorio_vendas.xlsx'

    wb.save(response)

    return response